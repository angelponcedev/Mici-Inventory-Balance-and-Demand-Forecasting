from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import math
from datetime import datetime, timedelta

def writeMissingData(file, forecastHorizonWeeks, quantityOfProducts):
    # All the forecasts will be written into a new file
    newFileAdress = file.replace(".xlsx"," WaferPlan + Forecasts.xlsx")
    # Loading base excel to work with
    wb = load_workbook(file)
    
    # loading sheets to work
    workSheetsList = [wb['Supply_Demand'],wb['Boundary Conditions'],wb['Wafer Plan']]
    
    #loading yield separetely cause of unique needs of writing columns
    Yield = wb['Yield']
    
    # Writing quarters for worksheets
    for workSheet in workSheetsList:
       writeColumns(ws=workSheet, forecastHorizonWeeks=forecastHorizonWeeks)
       
    #writing columns for yield
    writeYieldColumns(ws=Yield,forecastHorizonWeeks=forecastHorizonWeeks,quantityOfProducts=quantityOfProducts)
    
    #Filling the yield Sheet
    #yieldFilling(ws=Yield,forecastHorizonWeeks=forecastHorizonWeeks,quantityOfProducts=quantityOfProducts)
    
    # Saving the changes in a new workbook
    wb.save(newFileAdress)
    
    return newFileAdress

def writeColumns(ws,forecastHorizonWeeks):
    
    print('Proccesing ',ws.title)
    
    dimensions = ws.calculate_dimension()

    dimensions = dimensions.split(':')

    lastDimension = dimensions[1]

    col_part = ""
    row_part = ""
    split_index = -1

    for i, char in enumerate(lastDimension):
        # Check if the character is a digit
        if char.isdigit(): # Or: if char in string.digits:
            split_index = i
            break # Stop as soon as we find the first digit

    if split_index != -1:
        col_part = lastDimension[:split_index] # Slice from start up to (not including) the digit
        row_part = lastDimension[split_index:] # Slice from the digit to the end
    else:
        # Handle cases where there might be no numbers (though unlikely for cell refs)
        col_part = lastDimension
        row_part = "" # Or raise an error, depending on requirements

    print(f"Original: {lastDimension}")
    print(f"Column Part: {col_part}")
    print(f"Row Part: {row_part}")
    
    #Changing row part to 1 since that is the row where the quarters are
    row_part = 1

    lastQuarterString = ws[str(col_part)+str(row_part)].value

    print(lastQuarterString)

    numQuartersToWrite = math.ceil(forecastHorizonWeeks / 13)

    quartersToWrite = []

    quartersToWrite = generate_quarters(start_quarter=lastQuarterString, num_quarters=numQuartersToWrite)

    print(quartersToWrite)
    
    # 1. Get the integer index of the last written column
    try:
        last_col_index = column_index_from_string(col_part)
        # For "QB", last_col_index will be 444
    except ValueError:
        print(f"Error: Invalid column letters '{col_part}'")
        exit() # Handle error

    counter = 0
    for i in range(numQuartersToWrite):
        # Calculate the index for the *next* column
        # Add 1 to get the column *after* the last written one
        # Add 'i' for the current iteration (0 for the first new, 1 for second...)
        if ws.title == 'Boundary Conditions' or ws.title == 'Wafer Plan':
             timesToWrite = 14 # Write 13 times cause time is in weeks
        else:
            timesToWrite = 2 # write 1 time cause time is in quarters
        
        for weeks in range(1,timesToWrite):
            next_col_index = last_col_index + 1 + counter

            # Convert this index back to column letters (e.g., 445 -> "QC")
            next_col_letter = get_column_letter(next_col_index)

            # Calculate the next quarter string to write
            #current_quarter_str = lastQuarterString
            #current_quarter_str = get_next_quarter(current_quarter_str)
            current_quarter_str = quartersToWrite[i]
            if current_quarter_str is None:
                print("Stopping quarter generation due to an error.")
                break

            # Construct the target cell address (e.g., "QC1", "QD1", ...)
            target_cell = f"{next_col_letter}{row_part}"

            # Write the calculated quarter string to the target cell in the worksheet
            print(f"Writing '{current_quarter_str}' to cell {target_cell}")
            ws[target_cell] = current_quarter_str
            counter+=1

    if ws.title == 'Boundary Conditions' or ws.title == 'Wafer Plan':
        #--------------- Week handling -----------------------------------------------------
        #Changing row part of the extracted quarter string to 1 since thats the row where the week title are
        row_part = 2

        lastWeekString = ws[str(col_part)+str(row_part)].value

        print(lastWeekString)

        weeksToWrite = []

        weeksToWrite = generate_weeks(start_week=lastWeekString, num_weeks=forecastHorizonWeeks)

        print(weeksToWrite)
        
        # 1. Get the integer index of the last written column
        try:
            last_col_index = column_index_from_string(col_part)
            # For "QB", last_col_index will be 444
        except ValueError:
            print(f"Error: Invalid column letters '{col_part}'")
            exit() # Handle error

        for i in range(forecastHorizonWeeks):
            # Calculate the index for the *next* column
            # Add 1 to get the column *after* the last written one
            # Add 'i' for the current iteration (0 for the first new, 1 for second...)
            next_col_index = last_col_index + 1 + i

            # Convert this index back to column letters (e.g., 445 -> "QC")
            next_col_letter = get_column_letter(next_col_index)

            # Calculate the next quarter string to write
            #current_quarter_str = lastQuarterString
            #current_quarter_str = get_next_quarter(current_quarter_str)
            current_week_str = weeksToWrite[i]
            if current_week_str is None:
                print("Stopping quarter generation due to an error.")
                break

            # Construct the target cell address (e.g., "QC1", "QD1", ...)
            target_cell = f"{next_col_letter}{row_part}"

            # Write the calculated quarter string to the target cell in the worksheet
            print(f"Writing '{current_week_str}' to cell {target_cell}")
            ws[target_cell] = current_week_str
    return ws

def writeYieldColumns(ws,forecastHorizonWeeks,quantityOfProducts):
        
    print('Proccesing ',ws.title)
    
    dimensions = ws.calculate_dimension()

    dimensions = dimensions.split(':')

    lastDimension = dimensions[1]

    col_part = ""
    row_part = ""
    split_index = -1

    for i, char in enumerate(lastDimension):
        # Check if the character is a digit
        if char.isdigit(): # Or: if char in string.digits:
            split_index = i
            break # Stop as soon as we find the first digit

    if split_index != -1:
        col_part = lastDimension[:split_index] # Slice from start up to (not including) the digit
        row_part = lastDimension[split_index:] # Slice from the digit to the end
    else:
        # Handle cases where there might be no numbers (though unlikely for cell refs)
        col_part = lastDimension
        row_part = "" # Or raise an error, depending on requirements

    print(f"Original: {lastDimension}")
    print(f"Column Part: {col_part}")
    print(f"Row Part: {row_part}")
    
    col_index = column_index_from_string(col_part)
    
    col_index -= 12
    
    last_quarter_col_part = get_column_letter(col_index)
    
    #Changing row part to 1 since that is the row where the quarters are
    row_part = 1

    lastQuarterString = ws[str(last_quarter_col_part)+str(row_part)].value

    print(lastQuarterString)

    numQuartersToWrite = math.ceil(forecastHorizonWeeks / 13)
    
    print(numQuartersToWrite)

    quartersToWrite = []

    quartersToWrite = generate_quarters(start_quarter=lastQuarterString, num_quarters=numQuartersToWrite)

    print(quartersToWrite)
    
    # 1. Get the integer index of the last written column
    try:
        last_col_index = column_index_from_string(col_part)
        # For "QB", last_col_index will be 444
    except ValueError:
        print(f"Error: Invalid column letters '{col_part}'")
        exit() # Handle error

    counter = 0
    for i in range(numQuartersToWrite):
        # Calculate the index for the *next* column
        # Add 1 to get the column *after* the last written one
        # Add 'i' for the current iteration (0 for the first new, 1 for second...)

        next_col_index = last_col_index + 1 + counter * 13

        # Convert this index back to column letters (e.g., 445 -> "QC")
        next_col_letter = get_column_letter(next_col_index)

        # Calculate the next quarter string to write
        #current_quarter_str = lastQuarterString
        #current_quarter_str = get_next_quarter(current_quarter_str)
        current_quarter_str = quartersToWrite[i]
        if current_quarter_str is None:
            print("Stopping quarter generation due to an error.")
            break

        # Construct the target cell address (e.g., "QC1", "QD1", ...)
        target_cell = f"{next_col_letter}{row_part}"

        # Write the calculated quarter string to the target cell in the worksheet
        print(f"Writing '{current_quarter_str}' to cell {target_cell}")
        ws[target_cell] = current_quarter_str
        counter+=1
            
        
    #--------------- Week handling -----------------------------------------------------
    #Changing row part of the extracted quarter string to 1 since thats the row where the week title are
    row_part = 2

    lastWeekString = ws[str(col_part)+str(row_part)].value

    print(lastWeekString)

    weeksToWrite = []

    weeksToWrite = generate_dates(start_date_input=lastWeekString, num_dates=forecastHorizonWeeks)

    print(weeksToWrite)
    
    # 1. Get the integer index of the last written column
    try:
        last_col_index = column_index_from_string(col_part)
        # For "QB", last_col_index will be 444
    except ValueError:
        print(f"Error: Invalid column letters '{col_part}'")
        exit() # Handle error

    for i in range(forecastHorizonWeeks):
        # Calculate the index for the *next* column
        # Add 1 to get the column *after* the last written one
        # Add 'i' for the current iteration (0 for the first new, 1 for second...)
        next_col_index = last_col_index + 1 + i

        # Convert this index back to column letters (e.g., 445 -> "QC")
        next_col_letter = get_column_letter(next_col_index)

        # Calculate the next quarter string to write
        #current_quarter_str = lastQuarterString
        #current_quarter_str = get_next_quarter(current_quarter_str)
        current_week_str = weeksToWrite[i]
        if current_week_str is None:
            print("Stopping quarter generation due to an error.")
            break

        # Construct the target cell address (e.g., "QC1", "QD1", ...)
        target_cell = f"{next_col_letter}{row_part}"

        # Write the calculated quarter string to the target cell in the worksheet
        print(f"Writing '{current_week_str}' to cell {target_cell}")
        ws[target_cell] = current_week_str
        
    #----------- Yield per Product Filling ----------------------------------------------
    #Changing row part of the extracted quarter string to 1 since thats the row where the week title are
    for prod in range(quantityOfProducts):
        
        row_part = 3 + prod

        lastYield = ws[str(col_part)+str(row_part)].value

        print(lastYield)

        yieldsToWrite = []
        
        for i in range(forecastHorizonWeeks):
            yieldsToWrite.append(lastYield)

        print(yieldsToWrite)
        
        # 1. Get the integer index of the last written column
        try:
            last_col_index = column_index_from_string(col_part)
            # For "QB", last_col_index will be 444
        except ValueError:
            print(f"Error: Invalid column letters '{col_part}'")
            exit() # Handle error

        for i in range(forecastHorizonWeeks):
            # Calculate the index for the *next* column
            # Add 1 to get the column *after* the last written one
            # Add 'i' for the current iteration (0 for the first new, 1 for second...)
            next_col_index = last_col_index + 1 + i

            # Convert this index back to column letters (e.g., 445 -> "QC")
            next_col_letter = get_column_letter(next_col_index)

            # Calculate the next quarter string to write
            #current_quarter_str = lastQuarterString
            #current_quarter_str = get_next_quarter(current_quarter_str)
            current_week_yield = yieldsToWrite[i]
            if current_week_yield is None:
                print("Stopping yield clonning due to an error ocurring.")
                break

            # Construct the target cell address (e.g., "QC1", "QD1", ...)
            target_cell = f"{next_col_letter}{row_part}"

            # Write the calculated quarter string to the target cell in the worksheet
            print(f"Writing '{current_week_yield}' to cell {target_cell}")
            ws[target_cell] = current_week_yield
            
    return ws

def yieldFilling(ws,forecastHorizonWeeks,quantityOfProducts):
    print(f'{ws.title} filling')
    
    dimensions = ws.calculate_dimension()

    dimensions = dimensions.split(':')

    lastDimension = dimensions[1]

    col_part = ""
    row_part = ""
    split_index = -1

    for i, char in enumerate(lastDimension):
        # Check if the character is a digit
        if char.isdigit(): # Or: if char in string.digits:
            split_index = i
            break # Stop as soon as we find the first digit

    if split_index != -1:
        col_part = lastDimension[:split_index] # Slice from start up to (not including) the digit
        row_part = lastDimension[split_index:] # Slice from the digit to the end
    else:
        # Handle cases where there might be no numbers (though unlikely for cell refs)
        col_part = lastDimension
        row_part = "" # Or raise an error, depending on requirements

    print(f"Original: {lastDimension}")
    print(f"Column Part: {col_part}")
    print(f"Row Part: {row_part}")
        
    #--------------- Week handling -----------------------------------------------------
    #Changing row part of the extracted quarter string to 1 since thats the row where the week title are
    for prod in range(quantityOfProducts):
        
        row_part = 3 + prod

        lastYield = ws[str(col_part)+str(row_part)].value

        print(lastYield)

        yieldsToWrite = []
        
        for i in range(forecastHorizonWeeks):
            yieldsToWrite.append(lastYield)

        print(yieldsToWrite)
        
        # 1. Get the integer index of the last written column
        try:
            last_col_index = column_index_from_string(col_part)
            # For "QB", last_col_index will be 444
        except ValueError:
            print(f"Error: Invalid column letters '{col_part}'")
            exit() # Handle error

        for i in range(forecastHorizonWeeks):
            # Calculate the index for the *next* column
            # Add 1 to get the column *after* the last written one
            # Add 'i' for the current iteration (0 for the first new, 1 for second...)
            next_col_index = last_col_index + 1 + i

            # Convert this index back to column letters (e.g., 445 -> "QC")
            next_col_letter = get_column_letter(next_col_index)

            # Calculate the next quarter string to write
            #current_quarter_str = lastQuarterString
            #current_quarter_str = get_next_quarter(current_quarter_str)
            current_week_yield = yieldsToWrite[i]
            if current_week_yield is None:
                print("Stopping yield clonning due to an error ocurring.")
                break

            # Construct the target cell address (e.g., "QC1", "QD1", ...)
            target_cell = f"{next_col_letter}{row_part}"

            # Write the calculated quarter string to the target cell in the worksheet
            print(f"Writing '{current_week_yield}' to cell {target_cell}")
            ws[target_cell] = current_week_yield
            
    return ws


def generateSST(file, productsWithForecast):
    try:
        wb = load_workbook(file)
    except FileNotFoundError:
        print(f'Couldnt load File {file} as excel workbook')
        return
    except Exception as e:
        print(f'An exception ocurred while proccesing {file}, error: {e}')
        
    if 'wb' in locals() and wb is not None:    
        try:
            ws = wb['Supply_Demand']
            print("Successfully loaded 'Supply_Demand' sheet.")
            # You can now work with the 'ws' object
        except KeyError:
            # This is the specific error if the sheet name doesn't exist
            print(f"Error: Sheet 'Supply_Demand' not found in workbook '{file}'.")
            return
        
        except Exception as e:
            # Catch any other unexpected errors accessing the sheet
            print(f"Error accessing sheet 'Supply_Demand' in workbook '{file}': {e}")
            return
    else:
        print("Workbook was not loaded, cannot access sheet.")
            
    # In this point of the function the workbook and worksheet are loaded
    #getting the last dimension of the sheet
    print('Proccesing ',ws.title)
    
    dimensions = ws.calculate_dimension()

    dimensions = dimensions.split(':')

    lastDimension = dimensions[1]

    col_part = ""
    row_part = ""
    split_index = -1

    for i, char in enumerate(lastDimension):
        # Check if the character is a digit
        if char.isdigit(): # Or: if char in string.digits:
            split_index = i
            break # Stop as soon as we find the first digit

    if split_index != -1:
        col_part = lastDimension[:split_index] # Slice from start up to (not including) the digit
        row_part = lastDimension[split_index:] # Slice from the digit to the end
    else:
        # Handle cases where there might be no numbers (though unlikely for cell refs)
        col_part = lastDimension
        row_part = "" # Or raise an error, depending on requirements

    print(f"Original: {lastDimension}")
    print(f"Column Part: {col_part}")
    print(f"Row Part: {row_part}")
    
    lastDimensionInteger = column_index_from_string(col_part)
    
    # filling the sst for each product    
    for product in productsWithForecast:
        if product.productID == '21A':
            row_part = 3
        elif product.productID == '22B':
            row_part = 9
        elif product.productID == '23C':
            row_part = 15
                    
        demand_row = row_part + 2
        wos_row = row_part + 1
        weeks = 13
        starting_index = 3
        for i in range(starting_index,lastDimensionInteger):
            col_part = get_column_letter(i)
            next_col = get_column_letter(i+1)
            next_Demand = ws[next_col + str(demand_row)].value
            if math.isnan(next_Demand): 
                this_Quarter_Demand = ws[col_part + str(demand_row)].value
                ws[col_part + str(row_part)].value = (this_Quarter_Demand/weeks)*ws[col_part + str(wos_row)].value
            else:
                ws[col_part + str(row_part)].value = (next_Demand/weeks)*ws[col_part + str(wos_row)].value
            
        else:
            pass
        wb.save(file)
    return

def get_next_quarter(quarter_str):
    """
    Calculates the next quarter string in 'Qx yy' format.

    Args:
        quarter_str: The current quarter string (e.g., "Q4 03").

    Returns:
        The next quarter string (e.g., "Q1 04"), or None if input is invalid.
    """
    # Use regex to safely extract quarter and year
    match = re.match(r'^Q([1-4])\s+(\d{2})$', quarter_str)
    if not match:
        print(f"Invalid format: {quarter_str}")
        return None # Or raise an error

    q_num = int(match.group(1))
    year_yy = int(match.group(2))

    next_q_num = q_num + 1
    next_year_yy = year_yy

    if next_q_num > 4:
        next_q_num = 1
        next_year_yy += 1
        # Handle potential year 99 -> 00 rollover if needed,
        # though unlikely for recent fiscal years.
        if next_year_yy > 99:
             # Decide how to handle this - wrap to 00? Or go to 100?
             # Assuming wrap around for yy format:
             next_year_yy = 0 # Or maybe handle as an error/edge case

    # Format back to "Qx yy" with leading zero for the year
    return f"Q{next_q_num} {next_year_yy:02d}"

def generate_quarters(start_quarter, num_quarters):
    """Generates a list of subsequent quarter strings."""
    quarters = []
    current_quarter = start_quarter
    for _ in range(num_quarters):
        next_q = get_next_quarter(current_quarter)
        if next_q is None:
            print(f"Stopping generation due to invalid format from {current_quarter}")
            break # Stop if format becomes invalid
        quarters.append(next_q)
        current_quarter = next_q
    return quarters

def get_next_week(week_str):
    """
    Calculates the next week string in 'WW_nn' format (1-52 cycle).

    Args:
        week_str: The current week string (e.g., "WW_35").

    Returns:
        The next week string (e.g., "WW_36" or "WW_01" after "WW_52"),
        or None if input is invalid.
    """
    # Use regex to safely extract the week number. Allow WW_ or ww_.
    match = re.match(r'^WW_(\d{1,2})$', week_str, re.IGNORECASE)
    if not match:
        print(f"Invalid week format: '{week_str}'. Expected 'WW_nn'.")
        return None

    try:
        week_num = int(match.group(1))
    except ValueError:
         # This case is unlikely with \d+ regex but good practice
         print(f"Could not parse number in: '{week_str}'.")
         return None

    # Validate the week number is within the expected range (1-52)
    if not 1 <= week_num <= 52:
        print(f"Invalid week number: {week_num} in '{week_str}'. Must be between 1 and 52.")
        return None

    # Calculate the next week number with wrap-around
    if week_num == 52:
        next_week_num = 1
    else:
        next_week_num = week_num + 1

    # Format back to "WW_nn" with leading zero if needed (e.g., WW_01)
    return f"WW_{next_week_num:02d}"

def generate_weeks(start_week, num_weeks):
    """
    Generates a list of subsequent week strings based on the WW_nn format.

    Args:
        start_week: The starting week string (e.g., "WW_35").
        num_weeks: The number of subsequent weeks to generate.

    Returns:
        A list of generated week strings (e.g., ["WW_36", "WW_37", ...]).
        Returns an empty list if the start_week is invalid.
    """
    weeks = []
    current_week = start_week

    # --- Optional: Pre-validation of start_week ---
    # Check if the starting week format itself is valid before looping
    temp_match = re.match(r'^WW_(\d{1,2})$', start_week, re.IGNORECASE)
    if not temp_match:
         print(f"Invalid starting week format: '{start_week}'. Cannot generate weeks.")
         return []
    try:
        start_num = int(temp_match.group(1))
        if not 1 <= start_num <= 52:
             print(f"Invalid starting week number: {start_num} in '{start_week}'. Must be between 1 and 52.")
             return []
    except ValueError:
         print(f"Could not parse starting week number in: '{start_week}'.")
         return []
    # --- End Optional Pre-validation ---


    for _ in range(num_weeks):
        next_w = get_next_week(current_week)
        if next_w is None:
            # Error message already printed by get_next_week
            print(f"Stopping week generation due to invalid format derived from '{current_week}'.")
            break # Stop generation if an invalid state is reached
        weeks.append(next_w)
        current_week = next_w # Update for the next iteration

    return weeks


INPUT_DATE_FORMAT = "%d-%m-%Y"
OUTPUT_DATE_FORMAT = "%d-%m-%Y"

def get_next_date(date_input, days_to_add=7):
    """
    Calculates the date that is 'days_to_add' days after the given input.
    Handles input as either a 'YYYY-MM-DD' string (potentially with time)
    or a datetime.datetime object.
    Outputs in 'YYYY-MM-DD' format.

    Args:
        date_input: The current date (string or datetime.datetime object).
        days_to_add: The number of days to add (default is 7).

    Returns:
        The next date string in "YYYY-MM-DD" format,
        or None if the input is invalid.
    """
    current_date = None
    try:
        # Check if input is already a datetime object
        if isinstance(date_input, datetime):
            current_date = date_input
        # Check if input is a string
        elif isinstance(date_input, str):
            # Handle potential time component in string
            date_part = date_input.strip().split(' ')[0]
            # Parse the date part using the input format
            current_date = datetime.strptime(date_part, INPUT_DATE_FORMAT)
        else:
            # Input is neither a string nor a datetime object
            print(f"Invalid input type for date: {type(date_input)}. Expected string or datetime.")
            return None

        # Create a timedelta object representing the duration to add
        increment = timedelta(days=days_to_add)

        # Calculate the next date
        next_date = current_date + increment

        # Format the next date back into the desired output string format
        return next_date.strftime(OUTPUT_DATE_FORMAT)

    except ValueError:
        # Error during string parsing
        print(f"Invalid date format or value in string '{date_input}'. Expected date part as '{INPUT_DATE_FORMAT}'.")
        return None
    except TypeError:
         # Should be caught by the isinstance checks, but as a fallback
         print(f"Unexpected TypeError processing date input: {date_input}")
         return None
    except IndexError:
         # Handle cases where split might fail unexpectedly on a string
         print(f"Could not extract date part from string input: '{date_input}'.")
         return None


def generate_dates(start_date_input, num_dates, days_increment=7):
    """
    Generates a list of subsequent date strings, incrementing by a set number of days.
    Accepts start date as 'YYYY-MM-DD' string (optionally with time)
    or a datetime.datetime object.
    Outputs dates as 'YYYY-MM-DD'.

    Args:
        start_date_input: The starting date (string or datetime.datetime object).
        num_dates: The number of subsequent dates to generate.
        days_increment: The number of days between each generated date (default is 7).

    Returns:
        A list of generated date strings (e.g., ["2003-09-01", "2003-09-08", ...]).
        Returns an empty list if the start_date_input is invalid.
    """
    dates = []
    current_date_input = start_date_input # Can be string or datetime

    # --- Optional: Pre-validation of start_date_input ---
    is_valid_start = False
    if isinstance(start_date_input, datetime):
        is_valid_start = True # Datetime objects are inherently valid structurally
    elif isinstance(start_date_input, str):
        try:
            start_date_part = start_date_input.strip().split(' ')[0]
            datetime.strptime(start_date_part, INPUT_DATE_FORMAT)
            is_valid_start = True
        except (ValueError, TypeError, IndexError):
            is_valid_start = False # Error during parsing
    # else: input is neither string nor datetime

    if not is_valid_start:
        print(f"Invalid starting date input: '{start_date_input}' (Type: {type(start_date_input)}). Cannot generate dates.")
        return []
    # --- End Optional Pre-validation ---

    for i in range(num_dates):
        # Pass the current input (string or datetime) to get_next_date
        next_d_str = get_next_date(current_date_input, days_to_add=days_increment)

        if next_d_str is None:
            # Error message already printed by get_next_date
            print(f"Stopping date generation on iteration {i+1} due to invalid format derived from '{current_date_input}'.")
            break # Stop generation if an invalid state is reached

        dates.append(next_d_str)
        # IMPORTANT: Update current_date_input with the *string* output
        # for the next iteration, ensuring get_next_date receives
        # a consistent type after the first iteration.
        current_date_input = next_d_str

    return dates
