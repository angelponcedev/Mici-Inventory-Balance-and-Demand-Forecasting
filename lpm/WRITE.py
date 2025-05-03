# search.py (assuming mpl returns the result object)
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import csv 
import os

def write(result, producto, cuarto_inicio, weeks, output_path):
# --- Call the function and capture the result ---

    #output_path = r"D:/Universidad/TalentLand/LPM/Hackaton DB Final 04.21 W.xlsx"
    wafers = result  # Asegúrate de que esta variable tenga los datos correctos
    
    # --- CARGAR ARCHIVO ---
    book = load_workbook(output_path, data_only=True)
    sheet = book["Wafer Plan"]

    # --- DEFINIR FILA POR PRODUCTO ---
    if producto == "21A":
        start_row = 4
    elif producto == "22B":
        start_row = 5
    elif producto == "23C":
        start_row = 6
    else:
        raise ValueError(f"Producto no reconocido: {producto}")

    headers = [cell.value for cell in sheet[1]]  # Cambié la fila a 1 (fija que es la fila 2 en Excel)

    try:
        col_inicio = headers.index(cuarto_inicio) + 1  # Excel es 1-indexado
        #print(f"\nSe encontró '{cuarto_inicio}' en la columna {col_inicio}.")

        print("\nEscribiendo valores:")
        for i, value in enumerate(wafers):
            col_actual = col_inicio + i
            sheet.cell(row=start_row, column=col_actual, value=value)
            print(f"Celda ({start_row}, {col_actual}) ← {value}")

    except ValueError:
        print(f"\nERROR: No se encontró el cuarto '{cuarto_inicio}' en los encabezados.")
        print("Revisa si el texto coincide exactamente con lo que hay en la fila 1 del Excel (mayúsculas, espacios, etc).")
        
    book.save(output_path)
    print("\n💾 ¡Archivo guardado correctamente!")
    write2(result,producto,cuarto_inicio, weeks)
    return

def modify(yieldedSupply, tpib, ibesst, producto, cuarto, output_path):
    book = load_workbook(output_path, data_only=True)  # Preserva fórmulas
    sheet = book["Supply_Demand"]
    
    # Buscar filas dinámicamente
    product_rows = {}
    for row in sheet.iter_rows(min_row=2, max_row=19):
        product_id = row[0].value
        attribute = row[1].value
        if product_id == producto and attribute:
            if "Yielded Supply" in attribute:
                product_rows["yield"] = row[0].row
            elif "Total Projected Inventory Balance" in attribute:
                product_rows["tpib"] = row[0].row
            elif "Inventory Balance in excess of SST" in attribute:
                product_rows["ibesst"] = row[0].row
    
    # Validar que se encontraron las filas
    if not product_rows:
        raise KeyError(f"No se encontraron filas para el producto {producto}")
    
    # Buscar columna del trimestre (headers normalizados)
    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in sheet[1]]
    try:
        cuarto_buscar = str(cuarto).strip().upper()
        col_inicio = headers.index(cuarto_buscar) + 1
    except ValueError:
        print(f"Trimestre '{cuarto}' no encontrado. Encabezados disponibles: {headers}")
        return
    
    # Escribir valores (solo en celdas sin fórmulas)
    sheet.cell(row=product_rows["yield"], column=col_inicio, value=float(yieldedSupply))
    sheet.cell(row=product_rows["tpib"], column=col_inicio, value=float(tpib))
    sheet.cell(row=product_rows["ibesst"], column=col_inicio, value=float(ibesst))
    
    book.save(output_path)

def write2(result, producto, cuarto_inicio, weeks):
    nombre_archivo = "D:/Users/SEARS/Desktop/UAA/6toSemestre/TalentLand/Talentland-2025-main/Talentland-2025-main/dataset/waferPlan.csv"
    encabezados = ['Quarter', 'Week', 'Product', 'Wafers']
    wafers = result
    datos_nuevos = []

    for idx, val in enumerate(wafers):
        row = [
            cuarto_inicio,
            weeks[idx],
            producto,
            val
        ]
        datos_nuevos.append(row)

    # Si el archivo existe, leer y filtrar por Quarter Y Product
    if os.path.exists(nombre_archivo):
        with open(nombre_archivo, mode='r', newline='', encoding='utf-8') as archivo_csv:
            reader = csv.reader(archivo_csv)
            filas_existentes = list(reader)

        encabezado_leido = filas_existentes[0]
        cuerpo = filas_existentes[1:]

        # Filtrar filas que NO tengan el mismo cuarto Y producto
        cuerpo_filtrado = [
            fila for fila in cuerpo 
            if not (fila[0] == cuarto_inicio and fila[2] == producto)
        ]
    else:
        encabezado_leido = encabezados
        cuerpo_filtrado = []

    # Agregar los nuevos datos
    cuerpo_actualizado = cuerpo_filtrado + datos_nuevos

    # Sobrescribir el archivo completo con la combinación actualizada
    with open(nombre_archivo, mode='w', newline='', encoding='utf-8') as archivo_csv:
        writer = csv.writer(archivo_csv)
        writer.writerow(encabezado_leido)
        writer.writerows(cuerpo_actualizado)

    print(f"Datos de '{producto}' en el cuarto '{cuarto_inicio}' actualizados correctamente.")