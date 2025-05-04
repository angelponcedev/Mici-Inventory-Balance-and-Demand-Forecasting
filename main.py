import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Label, Text, Frame, Button, Canvas, Scrollbar, Radiobutton, Entry, Toplevel, IntVar, StringVar, BooleanVar
from PIL import Image, ImageTk
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates # Import for date formatting on charts
from openpyxl import load_workbook
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from dataset.dataset import processDataset
from forecast.forecast import generateCustomDemandForecast # Currently not implmented in the UI, Implement later
from forecast.Xgboost import generateDemandForecast
from boundaryPredictions.boundaryFactory import makeBoundaryPredictions 
import subprocess
from Mail_Window import AplicacionCorreo
from lpm.search import search
from lpm.search import search_combined
from excelHandling.writeMissingData import writeMissingData, generateSST
import os
import math

# --- Global Variables ---
supplyDemandSheet = None # Supply_Demand sheet
densityWafer = None # Density per Wafer
boundaryConditions = None
Yield = None
weeklyDemandRatio = None # Stores the weekly demand ratio per week
boundaries = None
forecastWeeks = None
cleanData = None # Stores the processed WEEKLY data suitable for forecasting
selectedProducts = [] # List of product IDs selected by the user
productsWithForecast = None # List of Product objects with WEEKLY forecast results
file = None

def guardar_predicciones_para_powerbi(productos_forecast, nombre_csv="predicciones_forecast.csv", nombre_pbix="plantilla_forecast.pbix"):
    try:
        datos = []

        for producto in productos_forecast:
            # Datos históricos
            if hasattr(producto, "quarter") and hasattr(producto, "demand"):
                for periodo, valor in zip(producto.quarter, producto.demand):
                    datos.append({
                        "Product ID": producto.productID,
                        "Periodo": periodo,
                        "Tipo": "Histórico",
                        "Valor Demanda": valor
                    })

            # Datos de predicción
            if hasattr(producto, "forecast"):
                for i, valor in enumerate(producto.forecast):
                    datos.append({
                        "Product ID": producto.productID,
                        "Periodo": f"F{i+1}",
                        "Tipo": "Pronóstico",
                        "Valor Demanda": valor
                    })

        df = pd.DataFrame(datos)
        df.to_csv(nombre_csv, index=False)
        messagebox.showinfo("Exportación Exitosa", f"Datos históricos y pronóstico guardados en {nombre_csv}")

        ruta_pbix = os.path.abspath(nombre_pbix)
        if os.path.exists(ruta_pbix):
            subprocess.Popen([ruta_pbix], shell=True)
        else:
            messagebox.showerror("Archivo no encontrado", f"No se encontró el archivo {nombre_pbix}")
    except Exception as e:
        messagebox.showerror("Error al exportar", f"Ocurrió un error: {e}")

# --- Utility Functions ---
# (Keep centerWindow, updateScrollregion, onMousewheel as they are)
def centerWindow(window, width, height):
    """Centers a Tkinter window on the screen."""
    screenWidth = window.winfo_screenwidth()
    screenHeight = window.winfo_screenheight()
    xPos = int((screenWidth - width) / 2)
    yPos = int((screenHeight - height) / 2)
    window.geometry(f'{width}x{height}+{xPos}+{yPos}')

def updateScrollregion(event=None):
    """Updates the scrollregion of the main canvas."""
    mainCanvas.configure(scrollregion=mainCanvas.bbox("all"))

def onMousewheel(event, widget):
    """Generic mousewheel scroll handler for scrollable widgets."""
    # Check for Linux vs Windows/macOS scroll events
    if event.num == 4 or event.delta > 0: # Scroll Up
        widget.yview_scroll(-1, "units")
    elif event.num == 5 or event.delta < 0: # Scroll Down
        widget.yview_scroll(1, "units")

# --- Core Application Logic Functions ---

def loadDataset():
    """Opens a file dialog to load a CSV/XLSX file."""
    global supplyDemandSheet, weeklyDemandRatio, boundaries, file
    file = filedialog.askopenfilename(
        title="Select Data File",
        filetypes=[("Excel files", "*.xlsx"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
    )
    if file:
        try:
            sheet_names = []
            if file.endswith('.xlsx'):
                xls = pd.ExcelFile(file)
                sheet_names = xls.sheet_names
                supplyDemandSheet = pd.read_excel(xls, sheet_name=sheet_names[0]) # Load first sheet as main data
                if 'Weekly Demand Ratio' in sheet_names:
                    weeklyDemandRatio = pd.read_excel(xls, sheet_name='Weekly Demand Ratio')
                else:
                    weeklyDemandRatio = None # Reset if sheet not found
                    print("Info: 'Weekly Demand Ratio' sheet not found in Excel file. Using default ratios if needed.")
                if 'Boundary Conditions' in sheet_names:
                    boundaries = pd.read_excel(xls, sheet_name='Boundary Conditions')
                else:
                    boundaries = None # Reset if sheet not found
                    print("Info: 'Boundary Conditions' sheet not found in Excel file. Using default boundaries if needed.")
            elif file.endswith('.csv'):
                supplyDemandSheet = pd.read_csv(file, delimiter=",", decimal=".")
                weeklyDemandRatio = None # Cannot load ratio from CSV easily
            else:
                 messagebox.showerror("Error", "Unsupported file type. Please select a CSV or XLSX file.")
                 return

            messagebox.showinfo("Success", f"Data loaded successfully:\n{file}")
            resetUIState() # Reset UI to initial state after loading new data
            resultText.insert(tk.END, f"Loaded {len(supplyDemandSheet)} rows from {file.split('/')[-1]}.\n")
            if weeklyDemandRatio is not None:
                 resultText.insert(tk.END, "Weekly demand ratios loaded from Excel sheet.\n")
            resultText.insert(tk.END, "Click 'Process Dataset' to prepare data for forecasting.\n")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load file: {str(e)}")
            supplyDemandSheet = None
            weeklyDemandRatio = None
            resetUIState()

def processDatasetAndShowSelector():
    """Processes the loaded dataset (converting to weekly) and opens the product selection window."""
    global supplyDemandSheet, cleanData, weeklyDemandRatio
    if supplyDemandSheet is None:
        messagebox.showwarning("Warning", "Please load a data file first.")
        return

    try:
        # Define default ratio if none was loaded
        currentRatio = weeklyDemandRatio
        if currentRatio is None:
           print("Using default weekly demand ratio.")
           # Ensure default is a list or compatible array-like structure
           currentRatio = [0.055,0.055,0.055,0.055,0.055,0.055,0.055,0.055,0.06,0.1,0.1,0.15,0.15]
           
        # Process dataset using the imported function - this now returns WEEKLY data
        cleanData = processDataset(supplyDemandSheet.copy(), currentRatio) # Pass ratio (or default)
        cleanData.to_csv('cleanDataQuarter.csv')
        

        if cleanData is None or cleanData.empty:
            messagebox.showinfo("Info", "No eligible products found or data processed after filtering.")
            resetUIState()
            return

        # Display summary in main window
        resultText.delete(1.0, tk.END)
        resultText.insert(tk.END, f"Dataset processed into weekly format.\n")
        resultText.insert(tk.END, f"Found {len(cleanData['Product ID'].unique())} eligible products.\n")
        resultText.insert(tk.END, "Opening product selection window...\n")

        # Show eligible products in a selection window
        showProductSelectionWindow()

    except Exception as e:
        messagebox.showerror("Error", f"Error processing dataset: {str(e)}")
        import traceback
        traceback.print_exc() # Print full traceback for debugging
        cleanData = None
        resetUIState()

def showProductSelectionWindow():
    """Displays a Toplevel window for selecting eligible products (using weekly data stats)."""
    global cleanData

    if cleanData is None or cleanData.empty:
        messagebox.showerror("Error", "No processed weekly data available to select products from.")
        return

    eligibleProductIDs = sorted(cleanData['Product ID'].unique().tolist())

    # --- Create Product Selection Window ---
    selectionWindow = Toplevel(window)
    selectionWindow.title("Product Selection for Forecasting")
    centerWindow(selectionWindow, 700, 550)
    selectionWindow.configure(bg="#F0F0F0")
    selectionWindow.grab_set()

    # Style configuration
    style = ttk.Style(selectionWindow)
    style.configure("TCheckbutton", background="#F0F0F0")
    style.configure("Treeview", rowheight=25, font=("Arial", 9))
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
    style.configure("TButton", padding=6, font=("Arial", 9))
    style.map("Treeview", background=[("selected", "#a6d1ff")])

    mainFrame = Frame(selectionWindow, bg="#F0F0F0")
    mainFrame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

    headerLabel = Label(mainFrame, text="Select products for weekly demand forecasting:",
                           font=("Arial", 11, "bold"), bg="#F0F0F0")
    headerLabel.pack(pady=(5, 10), anchor="w")

    contentFrame = Frame(mainFrame)
    contentFrame.pack(fill=tk.BOTH, expand=True)

    # Treeview setup
    tree = ttk.Treeview(contentFrame, columns=("select", "productID", "totalDemand", "numWeeks"), # Changed 'quarters' to 'numWeeks'
                        show="headings", height=15, style="Treeview")
    tree.heading("select", text="")
    tree.heading("productID", text="Product ID")
    tree.heading("totalDemand", text="Total Demand")
    tree.heading("numWeeks", text="Weeks of Data") # Changed heading

    tree.column("select", width=40, anchor=tk.CENTER, stretch=tk.NO)
    tree.column("productID", width=250, anchor=tk.W)
    tree.column("totalDemand", width=150, anchor=tk.E)
    tree.column("numWeeks", width=100, anchor=tk.CENTER) # Changed column

    scrollbar = ttk.Scrollbar(contentFrame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    tree.bind("<MouseWheel>", lambda e: onMousewheel(e, tree))
    tree.bind("<Button-4>", lambda e: onMousewheel(e, tree))
    tree.bind("<Button-5>", lambda e: onMousewheel(e, tree))

    # Calculate product statistics (using weekly data)
    productStats = {}
    for prodID in eligibleProductIDs:
        productData = cleanData[cleanData['Product ID'] == prodID]
        productStats[prodID] = {
            'total_demand': productData['Weekly Demand'].sum(),
            'num_weeks': len(productData) # Number of weekly data points
        }

    productVars = {prodID: BooleanVar(value=False) for prodID in eligibleProductIDs}
    tree.tag_configure('selected', background='#cce5ff')

    # Populate Treeview with weekly stats
    for prodID in eligibleProductIDs:
        stats = productStats[prodID]
        tree.insert("", tk.END, values=("☐", prodID, f"{stats['total_demand']:,.2f}", stats['num_weeks']), iid=prodID) # Use num_weeks

    # (Keep toggleItem, selectAll, deselectAll functions as they are)
    def toggleItem(event):
        item_id = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        if item_id and (column == '#1'):
            newValue = not productVars[item_id].get()
            productVars[item_id].set(newValue)
            values = list(tree.item(item_id, "values"))
            values[0] = "☑" if newValue else "☐"
            tree.item(item_id, values=tuple(values), tags=('selected',) if newValue else ())

    tree.bind("<Button-1>", toggleItem)

    buttonsFrame = Frame(mainFrame, bg="#F0F0F0")
    buttonsFrame.pack(pady=10, fill=tk.X)

    def selectAll():
        for prodID in eligibleProductIDs:
            if not productVars[prodID].get():
                productVars[prodID].set(True)
                values = list(tree.item(prodID, "values"))
                values[0] = "☑"
                tree.item(prodID, values=tuple(values), tags=('selected',))

    def deselectAll():
        for prodID in eligibleProductIDs:
            if productVars[prodID].get():
                productVars[prodID].set(False)
                values = list(tree.item(prodID, "values"))
                values[0] = "☐"
                tree.item(prodID, values=tuple(values), tags=())

    selectionButtonsFrame = Frame(buttonsFrame, bg="#F0F0F0")
    selectionButtonsFrame.pack(side=tk.LEFT, padx=5)
    btnSelectAll = ttk.Button(selectionButtonsFrame, text="Select All", command=selectAll, style="TButton")
    btnSelectAll.pack(side=tk.LEFT, padx=5)
    btnDeselectAll = ttk.Button(selectionButtonsFrame, text="Deselect All", command=deselectAll, style="TButton")
    btnDeselectAll.pack(side=tk.LEFT, padx=5)

    # Confirm Selection Button
    def confirmProductSelection():
        global selectedProducts
        selectedProducts = [prodID for prodID, var in productVars.items() if var.get()]
        if not selectedProducts:
            messagebox.showwarning("No Selection", "Please select at least one product.", parent=selectionWindow)
            return
        displaySelectedProductsInfo()
        selectionWindow.destroy()

    confirmFrame = Frame(buttonsFrame, bg="#F0F0F0")
    confirmFrame.pack(side=tk.RIGHT, padx=5)
    btnConfirm = Button(confirmFrame, text="Confirm Selection", command=confirmProductSelection,
                           bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                           padx=10, pady=5, relief=tk.RAISED, cursor="hand2")
    btnConfirm.pack()

    selectionWindow.wait_window()

def displaySelectedProductsInfo():
    """Updates the main data table with selected products (using weekly stats)."""
    global selectedProducts, cleanData, productsWithForecast

    dataTable.delete(*dataTable.get_children())
    productsWithForecast = None
    clearChartFrame("Select products and generate forecast to view chart.")

    # Configure data table columns for weekly view
    dataTable["columns"] = ("productID", "totalDemand", "avgWeeklyDemand", "numWeeks") # Adjusted columns
    dataTable.heading("productID", text="Product ID")
    dataTable.heading("totalDemand", text="Total Demand")
    dataTable.heading("avgWeeklyDemand", text="Avg. Weekly Demand") # Changed heading
    dataTable.heading("numWeeks", text="Weeks of Data") # Changed heading
    dataTable.column("#0", width=0, stretch=tk.NO)
    dataTable.column("productID", width=180, anchor=tk.W)
    dataTable.column("totalDemand", width=120, anchor=tk.E)
    dataTable.column("avgWeeklyDemand", width=120, anchor=tk.E) # Adjusted column
    dataTable.column("numWeeks", width=80, anchor=tk.CENTER) # Adjusted column

    # Populate the data table with selected product weekly stats
    if cleanData is not None and selectedProducts:
        for productID in selectedProducts:
            productData = cleanData[cleanData['Product ID'] == productID]
            if not productData.empty:
                totalDemand = productData['Weekly Demand'].sum()
                avgDemand = productData['Weekly Demand'].mean()
                numWeeks = len(productData) # Get number of weeks
                dataTable.insert("", "end", values=(productID, f"{totalDemand:,.2f}", f"{avgDemand:,.2f}", numWeeks), iid=productID) # Use weekly stats
            else:
                 dataTable.insert("", "end", values=(productID, "N/A", "N/A", "N/A"), iid=productID)

    # Update result text
    resultText.delete(1.0, tk.END)
    resultText.insert(tk.END, f"{len(selectedProducts)} products selected for weekly forecasting.\n\n")
    resultText.insert(tk.END, "Selected products:\n")
    for productID in selectedProducts:
        resultText.insert(tk.END, f"- {productID}\n")
    resultText.insert(tk.END, "\nClick 'Generate Forecast' to proceed.")

    btnGenerateForecast.config(state=tk.NORMAL if selectedProducts else tk.DISABLED)
    dataTable.unbind("<<TreeviewSelect>>")

def selectForecastPeriod():
    """Opens a modal window to select the number of WEEKS for forecasting."""
    periodWindow = Toplevel(window)
    periodWindow.title("Forecast Horizon") # Changed title
    centerWindow(periodWindow, 350, 180)
    periodWindow.configure(bg="#F5F5F5")
    periodWindow.resizable(False, False)
    periodWindow.grab_set()

    periodWindow.selected_weeks = None # Changed attribute name

    mainFrame = Frame(periodWindow, bg="#F5F5F5")
    mainFrame.pack(pady=15, padx=20, fill=tk.BOTH, expand=True)

    Label(mainFrame, text="Select number of Weeks to Forecast:", # Changed text
             font=("Arial", 10), bg="#F5F5F5").pack(pady=(5, 10))

    periodVar = IntVar(value=13) # Default to 13 weeks (1 quarter)
    # Provide weekly options (e.g., 1 quarter to 1 year)
    week_options = [13, 26, 39, 52, 65, 78, 91, 104, 117, 130, 143, 156]
    combo = ttk.Combobox(mainFrame, textvariable=periodVar, state="readonly",
                         values=week_options, width=10, font=("Arial", 10))
    combo.pack(pady=5)
    # Find index of default value 13, default to 0 if not found
    try:
        default_index = week_options.index(13)
        combo.current(default_index)
    except ValueError:
        combo.current(0)


    buttonFrame = Frame(mainFrame, bg="#F5F5F5")
    buttonFrame.pack(pady=15)

    def confirm():
        periodWindow.selected_weeks = periodVar.get() # Store selected weeks
        periodWindow.grab_release()
        periodWindow.destroy()

    def cancel():
        periodWindow.selected_weeks = None
        periodWindow.grab_release()
        periodWindow.destroy()

    confirmBtn = Button(buttonFrame, text="Confirm", command=confirm,
                           bg="#4CAF50", fg="white", font=("Arial", 9, "bold"),
                           width=8, relief=tk.RAISED, cursor="hand2")
    confirmBtn.pack(side=tk.LEFT, padx=10)

    cancelBtn = Button(buttonFrame, text="Cancel", command=cancel,
                          bg="#e74c3c", fg="white", font=("Arial", 9, "bold"),
                          width=8, relief=tk.RAISED, cursor="hand2")
    cancelBtn.pack(side=tk.LEFT, padx=10)

    periodWindow.protocol("WM_DELETE_WINDOW", cancel)
    periodWindow.wait_window()

    return periodWindow.selected_weeks # Return selected weeks

def callForecasting():
    global productsWithForecast, file
    
    generateForecastAndDisplay()
    generateBoundaryForecast()
    for product in productsWithForecast:
         if hasattr(product, 'productID') and hasattr(product, 'boundaries'):
             boundaries_str = str(product.boundaries[:10]) + ('...' if len(product.boundaries) > 10 else '')
             mape_str = f"{product.backtestMetrics.get('mape', 'N/A'):.2f}%" if isinstance(product.backtestMetrics.get('mape'), (int, float)) else product.backtestMetrics.get('error', 'N/A')
             print(f'Product {product.productID} boundaries ({len(product.boundaries)} total): {boundaries_str}')
             print(f'  Backtest MAPE: {mape_str}\n')

    # Writing SST before doing the LPM proccesing
    generateSST(file=file, productsWithForecast=productsWithForecast)

    quarters = math.ceil(forecastWeeks / 13)
    years = quarters // 4      # División entera para obtener los años completos
    quarter = quarters % 4
    #search_combined(file,years+9,quarter)
    search(file,years+9,quarter)
    
def generateBoundaryForecast():
    """
    Generates boundary forecast for products and writes predictions to an Excel file.
    Uses global variables for product data, forecast settings, and file paths.
    """
    global boundaries, productsWithForecast, file, boundaryConditions, forecastWeeks
    
    # Validate that required forecast data is available
    if productsWithForecast is None or forecastWeeks is None:
        messagebox.showerror(
            "Error", 
            "Boundary prediction needs a list of products to operate, make a forecast for products beforehand."
        )
        return
    
    # Generate boundary predictions for all products with forecast
    makeBoundaryPredictions(boundaries, productsWithForecast, forecastWeeks)
    
    # Write predictions to Excel file for each product
    for product in productsWithForecast:
        # Create a list of boundary indices for the forecast weeks
        boundaryList = []
        boundaryList = product.boundaries[-forecastWeeks:]
        
        # Write the predictions to the Excel file
        writeBoundaryPredictions(
            predictions=boundaryList,
            producto=product.productID,
            cuarto_inicio='Q1 04',
            file=file
        )
        
    return    

def writeForecastPredictions(predictions, producto, cuarto_inicio, file):
    """
    Writes forecast predictions to an Excel file.
    
    Args:
        predictions (list): List of boundary values to write
        producto (str): Product ID (21A, 22B, or 23C)
        cuarto_inicio (str): Starting quarter column header (e.g., 'Q1 04')
        file (str): Path to the Excel file
    
    Returns:
        None
    """
    # Set output path
    output_path = rf"{file}"
    availableCapacity = predictions
    
    # Load the Excel workbook
    book = load_workbook(output_path, data_only=True)
    sheet = book["Supply_Demand"]

    # Determine the starting row based on product ID
    if producto == "21A":
        start_row = 5
    elif producto == "22B":
        start_row = 11
    elif producto == "23C":
        start_row = 17
    else:
        raise ValueError(f"Producto no reconocido: {producto}")

    # Get the header row values from the first row
    headers = [cell.value for cell in sheet[1]]  # First row in Excel (index 1)

    try:
        # Find the starting column based on the quarter header
        col_inicio = headers.index(cuarto_inicio) + 1  # Excel uses 1-based indexing
        print(f"\nSe encontró '{cuarto_inicio}' en la columna {col_inicio}.")

        # Write each prediction value to the appropriate cell
        print("\nEscribiendo valores:")
        for i, value in enumerate(availableCapacity):
            col_actual = col_inicio + i
            sheet.cell(row=start_row, column=col_actual, value=value)
            #print(f"Celda ({start_row}, {col_actual}) ← {value}")

    except ValueError:
        # Handle case where the quarter header is not found
        print(f"\nERROR: No se encontró el cuarto '{cuarto_inicio}' en los encabezados.")
        print("Revisa si el texto coincide exactamente con lo que hay en la fila 1 del Excel (mayúsculas, espacios, etc).")
    
    # Save the workbook
    book.save(output_path)
    print("\n💾 ¡Archivo guardado correctamente!")
    return

def writeWOS(predictions, producto, cuarto_inicio, file):
    """
    Writes forecast predictions to an Excel file.
    
    Args:
        predictions (list): List of boundary values to write
        producto (str): Product ID (21A, 22B, or 23C)
        cuarto_inicio (str): Starting quarter column header (e.g., 'Q1 04')
        file (str): Path to the Excel file
    
    Returns:
        None
    """
    # Set output path
    output_path = rf"{file}"
    
    wosList = []
    
    for i in range(len(predictions)):
        if producto == '21A':
            wosList.append(1.82)
        if producto == '22B':
            wosList.append(1.56)
        if producto == '23C':
            wosList.append(1.42)
    
    # Load the Excel workbook
    book = load_workbook(output_path, data_only=True)
    sheet = book["Supply_Demand"]

    # Determine the starting row based on product ID
    if producto == "21A":
        start_row = 4
    elif producto == "22B":
        start_row = 10
    elif producto == "23C":
        start_row = 16
    else:
        raise ValueError(f"Producto no reconocido: {producto}")

    # Get the header row values from the first row
    headers = [cell.value for cell in sheet[1]]  # First row in Excel (index 1)

    try:
        # Find the starting column based on the quarter header
        col_inicio = headers.index(cuarto_inicio) + 1  # Excel uses 1-based indexing
        print(f"\nSe encontró '{cuarto_inicio}' en la columna {col_inicio}.")

        # Write each prediction value to the appropriate cell
        print("\nEscribiendo valores:")
        for i, value in enumerate(wosList):
            col_actual = col_inicio + i
            sheet.cell(row=start_row, column=col_actual, value=value)
            #print(f"Celda ({start_row}, {col_actual}) ← {value}")

    except ValueError:
        # Handle case where the quarter header is not found
        print(f"\nERROR: No se encontró el cuarto '{cuarto_inicio}' en los encabezados.")
        print("Revisa si el texto coincide exactamente con lo que hay en la fila 1 del Excel (mayúsculas, espacios, etc).")
    
    # Save the workbook
    book.save(output_path)
    print("\n💾 ¡Archivo guardado correctamente!")
    return


def writeBoundaryPredictions(predictions, producto, cuarto_inicio, file):
    """
    Writes boundary predictions to an Excel file.
    
    Args:
        predictions (list): List of boundary values to write
        producto (str): Product ID (21A, 22B, or 23C)
        cuarto_inicio (str): Starting quarter column header (e.g., 'Q1 04')
        file (str): Path to the Excel file
    
    Returns:
        None
    """
    # Set output path
    output_path = rf"{file}"
    availableCapacity = predictions
    
    # Load the Excel workbook
    book = load_workbook(output_path, data_only=True)
    sheet = book["Boundary Conditions"]

    # Determine the starting row based on product ID
    if producto == "21A":
        start_row = 6
    elif producto == "22B":
        start_row = 9
    elif producto == "23C":
        start_row = 12
    else:
        raise ValueError(f"Producto no reconocido: {producto}")

    # Get the header row values from the first row
    headers = [cell.value for cell in sheet[1]]  # First row in Excel (index 1)

    try:
        # Find the starting column based on the quarter header
        col_inicio = headers.index(cuarto_inicio) + 1  # Excel uses 1-based indexing
        print(f"\nSe encontró '{cuarto_inicio}' en la columna {col_inicio}.")

        # Write each prediction value to the appropriate cell
        print("\nEscribiendo valores:")
        for i, value in enumerate(availableCapacity):
            col_actual = col_inicio + i
            sheet.cell(row=start_row, column=col_actual, value=value)
            #print(f"Celda ({start_row}, {col_actual}) ← {value}")

    except ValueError:
        # Handle case where the quarter header is not found
        print(f"\nERROR: No se encontró el cuarto '{cuarto_inicio}' en los encabezados.")
        print("Revisa si el texto coincide exactamente con lo que hay en la fila 1 del Excel (mayúsculas, espacios, etc).")
    
    # Save the workbook
    book.save(output_path)
    print("\n💾 ¡Archivo guardado correctamente!")
    return

def generateForecastAndDisplay():
    """Generates WEEKLY forecast and displays the first chart."""
    global selectedProducts, cleanData, productsWithForecast, forecastWeeks, file

    if not selectedProducts:
        messagebox.showwarning("Warning", "Please select products first.")
        return
    if cleanData is None:
         messagebox.showerror("Error", "Cannot forecast without processed weekly data.")
         return

    forecastWeeks = selectForecastPeriod() # Get number of weeks
    if forecastWeeks is None:
        resultText.insert(tk.END, "\nForecast cancelled by user.\n")
        return
    
    #------ Adding columns to the file in accordance to forecastWeeks --------------------------
    # Genearating column names for the file before passing it
    # important this function returns a new filepath, this means the path inside file is changed to the output file
    file = writeMissingData(file=file,forecastHorizonWeeks=forecastWeeks,quantityOfProducts=len(selectedProducts))
    
    #-------------------------------------------------------------------------------------------

    try:
        resultText.insert(tk.END, f"\nGenerating forecast for {forecastWeeks} weeks...\n") # Update text
        window.update_idletasks()

        # Call the weekly forecast function from forecast.py
        # Pass the list of product IDs, number of weeks, and the weekly cleanData
        productsWithForecast = generateDemandForecast(
            productIDs=selectedProducts,
            forecastHorizonWeeks=forecastWeeks,
            historicalWeeklyData=cleanData
        )
        
        for product in productsWithForecast:
            # Calculate quarterly demand by summing groups of 13 weekly predictions
            quartersDemand = []
            
            # Determine how many complete quarters we can make
            num_quarters = forecastWeeks // 13
            
            # Process each complete quarter (13 weeks per quarter)
            for q in range(num_quarters):
                start_idx = q * 13
                end_idx = start_idx + 13
                # Sum the next 13 weekly predictions to create one quarterly value
                quarter_sum = sum(product.forecast[start_idx:end_idx])
                quartersDemand.append(quarter_sum)
                
            # Handle any remaining weeks (partial quarter)
            remaining_weeks = forecastWeeks % 13
            if remaining_weeks > 0:
                start_idx = num_quarters * 13
                partial_quarter_sum = sum(product.forecast[start_idx:])
                quartersDemand.append(partial_quarter_sum)
            
            # Write quarterly demand predictions to file
            writeForecastPredictions(
                predictions=quartersDemand,
                producto=product.productID,
                cuarto_inicio='Q1 04',
                file=file
            )
            
            # Write wos constant to file for each product
            writeWOS(
                predictions=quartersDemand,
                producto=product.productID,
                cuarto_inicio='Q1 04',
                file=file
            )

        if not productsWithForecast:
             messagebox.showerror("Error", "Weekly forecasting failed or returned no results.")
             clearChartFrame("Forecasting Error.")
             resultText.insert(tk.END, "Weekly forecasting failed.\n")
             # Ensure binding is removed on error too
             dataTable.unbind("<<TreeviewSelect>>")
             btnCalculateBalance.config(state=tk.DISABLED) # Disable balance calc
             return

        messagebox.showinfo("Success", f"Weekly forecast generated for {forecastWeeks} weeks.")
        resultText.insert(tk.END, f"Weekly forecast generated successfully for {len(productsWithForecast)} products.\n")
        resultText.insert(tk.END, "Click on a product in the table to view its demand chart.\n")

        # Enable balance calculation button now that forecast exists
        if lpmOptionVar.get() == "balance":
             btnCalculateBalance.config(state=tk.NORMAL)

        # Display the chart for the first selected product
        if selectedProducts:
            firstProductID = selectedProducts[0]
            if firstProductID in dataTable.get_children():
                 dataTable.selection_set(firstProductID)
                 dataTable.focus(firstProductID)
                 displayDemandChart(firstProductID) # Display chart immediately
            else:
                 clearChartFrame(f"Data for {firstProductID} not found in table.")

            # Bind Treeview selection to update chart AFTER forecast
            dataTable.unbind("<<TreeviewSelect>>")
            dataTable.bind("<<TreeviewSelect>>", onProductSelectInTable)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred during weekly forecasting: {str(e)}")
        import traceback
        traceback.print_exc() # Print full traceback
        productsWithForecast = None
        clearChartFrame("Error during forecast generation.")
        resultText.insert(tk.END, f"\nWeekly forecasting error: {e}\n")
        dataTable.unbind("<<TreeviewSelect>>")
        btnCalculateBalance.config(state=tk.DISABLED) # Disable balance calc

# (Keep onProductSelectInTable as is)
def onProductSelectInTable(event):
    """Handles product selection in the table to update the chart."""
    if productsWithForecast is None:
        return
    selectedItems = dataTable.selection()
    if selectedItems:
        productID = selectedItems[0]
        displayDemandChart(productID)

# (Keep clearChartFrame as is)
def clearChartFrame(message=""):
    """Clears the chart frame and displays an optional placeholder message."""
    for widget in chartDisplayFrame.winfo_children():
        widget.destroy()
    if message:
        placeholderLabel = Label(chartDisplayFrame, text=message,
                                 font=("Arial", 10), bg="white", fg="#555")
        placeholderLabel.pack(pady=20, padx=10)

def displayDemandChart(productID):
    """Displays the WEEKLY historical demand and forecast chart."""
    clearChartFrame()

    if productsWithForecast is None:
        clearChartFrame("Generate a forecast to view charts.")
        return

    productObject = next((p for p in productsWithForecast if p.productID == productID), None)

    if productObject is None:
        clearChartFrame(f"Forecast data not found for product: {productID}")
        return

    try:
        fig = plt.Figure(figsize=(6, 4), dpi=90)
        ax = fig.add_subplot(111)
        ax.set_facecolor('#FEFEFE')
        fig.patch.set_facecolor('white')

        lastHistoricalDate = None
        lastHistoricalValue = None

        # --- Plot Historical Weekly Data ---
        if hasattr(productObject, 'dates') and hasattr(productObject, 'demand') and \
           productObject.dates and productObject.demand and \
           len(productObject.dates) == len(productObject.demand):

            # Filter out potential NaT dates if any exist
            valid_hist_indices = [i for i, d in enumerate(productObject.dates) if pd.notna(d)]
            historicalDates = [productObject.dates[i] for i in valid_hist_indices]
            historicalDemand = [productObject.demand[i] for i in valid_hist_indices]

            if historicalDates: # Check if list is not empty after filtering
                ax.plot(historicalDates, historicalDemand, marker='.', linestyle='-',
                        color='#1f77b4', label='Historical Demand', markersize=4) # Smaller marker
                lastHistoricalDate = historicalDates[-1]
                lastHistoricalValue = historicalDemand[-1]
            else:
                 print(f"Warning: No valid historical dates found for {productID}")

        else:
            print(f"Warning: Missing or mismatched historical weekly data for {productID}")

        # --- Plot Forecast Weekly Data ---
        if hasattr(productObject, 'forecastDates') and hasattr(productObject, 'forecast') and \
           productObject.forecastDates and productObject.forecast:

            # Filter out potential NaT dates
            valid_fc_indices = [i for i, d in enumerate(productObject.forecastDates) if pd.notna(d)]
            forecastDates = [productObject.forecastDates[i] for i in valid_fc_indices]
            forecastValues = [productObject.forecast[i] for i in valid_fc_indices]


            if forecastDates: # Check if list is not empty
                plotDates = forecastDates
                plotValues = forecastValues
                # Connect forecast line to last historical point if available
                if lastHistoricalDate and lastHistoricalValue is not None:
                    plotDates = [lastHistoricalDate] + forecastDates
                    plotValues = [lastHistoricalValue] + forecastValues

                ax.plot(plotDates, plotValues, marker='.', linestyle='--',
                        color='#ff7f0e', label='Forecast', markersize=4) # Smaller marker
            else:
                print(f"Warning: No valid forecast dates found for {productID}")

        # --- Configure Chart Appearance ---
        ax.set_title(f"Weekly Demand Analysis: {productID}", fontsize=12, weight='bold')
        ax.set_xlabel("Date", fontsize=10)
        ax.set_ylabel("Demand", fontsize=10)

        # Improve x-axis tick display for dates
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=4, maxticks=12)) # Auto ticks
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d')) # Format ticks
        fig.autofmt_xdate(rotation=30, ha='right') # Auto-rotate date labels

        ax.grid(True, linestyle=':', linewidth=0.6, color='gray', alpha=0.7)
        ax.legend(loc='best', fontsize=9)
        fig.tight_layout(pad=0.5) # Adjust layout padding

        # Embed Chart in Tkinter Frame
        canvas = FigureCanvasTkAgg(fig, master=chartDisplayFrame)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill=tk.BOTH, expand=True)

    except Exception as e:
        messagebox.showerror("Chart Error", f"Could not display weekly chart for {productID}: {str(e)}")
        import traceback
        traceback.print_exc()
        clearChartFrame(f"Error displaying chart for {productID}.")

def abrir_correo():
    nueva_ventana = Toplevel(window)
    AplicacionCorreo(nueva_ventana)

def resetUIState():
    """Resets the UI elements to their initial state."""
    global selectedProducts, productsWithForecast, cleanData, supplyDemandSheet
    

    resultText.delete(1.0, tk.END)
    resultText.insert(tk.END, "Load a dataset (CSV/XLSX) to begin.\n")
    clearChartFrame("Load data, process, and select products to view charts.")

    # Clear and reconfigure table for weekly view
    dataTable.delete(*dataTable.get_children())
    dataTable["columns"] = ("productID", "totalDemand", "avgWeeklyDemand", "numWeeks")
    dataTable.heading("productID", text="Product ID")
    dataTable.heading("totalDemand", text="Total Demand")
    dataTable.heading("avgWeeklyDemand", text="Avg. Weekly Demand")
    dataTable.heading("numWeeks", text="Weeks of Data")
    dataTable.column("#0", width=0, stretch=tk.NO)
    dataTable.column("productID", width=180, anchor=tk.W)
    dataTable.column("totalDemand", width=120, anchor=tk.E)
    dataTable.column("avgWeeklyDemand", width=120, anchor=tk.E)
    dataTable.column("numWeeks", width=80, anchor=tk.CENTER)

    selectedProducts = []
    productsWithForecast = None
    cleanData = None

    # Reset LPM inputs if they exist
    if 'initialStockEntry' in globals() and initialStockEntry.winfo_exists():
        initialStockEntry.delete(0, tk.END)
    if 'receiptsEntry' in globals() and receiptsEntry.winfo_exists():
        receiptsEntry.delete(0, tk.END)
    if 'safetyStockEntry' in globals() and safetyStockEntry.winfo_exists():
        safetyStockEntry.delete(0, tk.END)
    if 'variable4Entry' in globals() and variable4Entry.winfo_exists():
        variable4Entry.delete(0, tk.END)
    if 'variable5Entry' in globals() and variable5Entry.winfo_exists():
        variable5Entry.delete(0, tk.END)
    if 'variable6Entry' in globals() and variable6Entry.winfo_exists():
        variable6Entry.delete(0, tk.END)
    if 'balanceResultLabel' in globals() and balanceResultLabel.winfo_exists():
        balanceResultLabel.config(text="Final Demand: N/A")
    if 'btnCalculateBalance' in globals() and btnCalculateBalance.winfo_exists():
        btnCalculateBalance.config(state=tk.DISABLED)


    btnProcessDataset.config(state=tk.NORMAL if supplyDemandSheet is not None else tk.DISABLED)
    btnGenerateForecast.config(state=tk.DISABLED)
    # btnCalculateBalance state is handled within updateLpmPanel and generateForecast

    dataTable.unbind("<<TreeviewSelect>>")


# --- LPM Solution Section Functions ---
def createLabeledEntry(parent, labelText, entryVar=None):
    """Helper to create a Label and Entry pair."""
    frame = Frame(parent, bg="white")
    frame.pack(pady=3, padx=5, fill=tk.X)
    label = Label(frame, text=labelText, width=12, anchor="w", bg="white", font=("Arial", 9))
    label.pack(side=tk.LEFT, padx=(0, 5))
    entry = Entry(frame, font=("Arial", 9), bd=1, relief=tk.SOLID, textvariable=entryVar)
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
    return entry

# (Keep updateLpmPanel as is, it correctly enables/disables btnCalculateBalance)
def updateLpmPanel():
    """Updates the content of the right LPM panel based on radio button selection."""
    for widget in lpmInputFrame.winfo_children():
        widget.destroy()

    global initialStockEntry, receiptsEntry, safetyStockEntry
    global variable4Entry, variable5Entry, variable6Entry
    global balanceResultLabel, btnCalculateBalance

    selectedOption = lpmOptionVar.get()

    if selectedOption == "balance":
        initialStockEntry = createLabeledEntry(lpmInputFrame, "Initial Stock:")
        receiptsEntry = createLabeledEntry(lpmInputFrame, "Receipts:")
        safetyStockEntry = createLabeledEntry(lpmInputFrame, "Safety Stock:")
        variable4Entry = createLabeledEntry(lpmInputFrame, "Variable 4:")
        variable5Entry = createLabeledEntry(lpmInputFrame, "Variable 5:")
        variable6Entry = createLabeledEntry(lpmInputFrame, "Variable 6:")

        calcFrame = Frame(lpmInputFrame, bg="white")
        calcFrame.pack(pady=10, fill=tk.X)

        btnCalculateBalance = Button(calcFrame, text="Calculate Balance", command=calculateBalance,
                                     bg="#f39c12", fg="white", font=("Arial", 10, "bold"),
                                     padx=10, pady=5, cursor="hand2", state=tk.DISABLED)
        btnCalculateBalance.pack(pady=5)

        balanceResultLabel = Label(calcFrame, text="Final Demand: N/A",
                                   bg="white", font=("Arial", 10, "bold"), fg="#333")
        balanceResultLabel.pack(pady=5)

        if productsWithForecast: # Enable if forecast exists
             btnCalculateBalance.config(state=tk.NORMAL)

    elif selectedOption == "option2":
        Label(lpmInputFrame, text="Option 2 Content Area", bg="white", font=("Arial", 10)).pack(pady=20)
    elif selectedOption == "option3":
        Label(lpmInputFrame, text="Exportar a Power BI", bg="white", font=("Arial", 10)).pack(pady=20)
        
        def ejecutarExportacion():
            if productsWithForecast:
                guardar_predicciones_para_powerbi(productsWithForecast)
            else:
                messagebox.showwarning("Advertencia", "Primero debes generar una predicción.")

        btnExportarPBIX = Button(lpmInputFrame, text="Exportar y Abrir Power BI",
                                command=ejecutarExportacion,
                                bg="#007ACC", fg="white", font=("Arial", 10, "bold"),
                                padx=10, pady=5, cursor="hand2")
        btnExportarPBIX.pack(pady=10)


def calculateBalance():
    """Calculates a balance based on inputs and SUM of weekly forecast data."""
    global productsWithForecast

    
    selectedItems = dataTable.selection()
    if not selectedItems:
        messagebox.showwarning("Warning", "Please select a product in the table to calculate balance for.")
        return
    if productsWithForecast is None:
        messagebox.showwarning("Warning", "Forecast data is not available. Please generate a forecast first.")
        return

    productID = selectedItems[0]
    productObject = next((p for p in productsWithForecast if p.productID == productID), None)

    if productObject is None or not hasattr(productObject, 'forecast') or not productObject.forecast:
        messagebox.showerror("Error", f"Weekly forecast data for product {productID} is missing or invalid.")
        return

    # Use the SUM of the weekly forecast values over the horizon
    try:
        # Ensure forecast list contains numbers before summing
        numeric_forecast = [f for f in productObject.forecast if isinstance(f, (int, float))]
        if len(numeric_forecast) != len(productObject.forecast):
             print(f"Warning: Non-numeric values found in forecast for {productID}. Summing only numeric values.")
        if not numeric_forecast:
             raise ValueError("Forecast list contains no numeric values.")
        forecastDemand = sum(numeric_forecast)
    except TypeError:
        messagebox.showerror("Error", f"Forecast data for product {productID} contains non-numeric types.")
        return
    except ValueError as ve:
         messagebox.showerror("Error", f"Could not process forecast demand for {productID}: {ve}")
         return
    except Exception as e:
         messagebox.showerror("Error", f"Could not process forecast demand for {productID}: {str(e)}")
         return

    try:
        initial = float(initialStockEntry.get() or 0)
        receipts = float(receiptsEntry.get() or 0)
        safeStock = float(safetyStockEntry.get() or 0)
        # var4 = float(variable4Entry.get() or 0) # Uncomment if used
        # var5 = float(variable5Entry.get() or 0) # Uncomment if used
        # var6 = float(variable6Entry.get() or 0) # Uncomment if used

        # Example Balance Calculation
        balanceResult = initial + receipts - forecastDemand - safeStock

        balanceResultLabel.config(text=f"Final Demand ({productID}): {balanceResult:,.2f}")

    except ValueError:
        messagebox.showerror("Input Error", "Please enter valid numbers in all required fields.")
    except Exception as e:
         messagebox.showerror("Calculation Error", f"An error occurred during calculation: {str(e)}")


# ==============================================================================
# --- Main Application Window Setup (Keep the rest of the GUI layout code) ---
# ==============================================================================

window = tk.Tk()
window.title("MICI Demand Forecasting v2.1 - Weekly") # Updated Title
window.state('zoomed')
window.configure(bg="#EAEAEA")

# --- Main Canvas and Scrollbar ---
mainCanvas = Canvas(window, bg="#EAEAEA", highlightthickness=0)
mainScrollbar = ttk.Scrollbar(window, orient="vertical", command=mainCanvas.yview)
mainCanvas.configure(yscrollcommand=mainScrollbar.set)
mainScrollbar.pack(side="right", fill="y")
mainCanvas.pack(side="left", fill="both", expand=True)
contentFrame = Frame(mainCanvas, bg="#030075")
mainCanvas.create_window((0, 0), window=contentFrame, anchor="nw", tags="content_frame")
mainCanvas.bind("<Configure>", lambda e: mainCanvas.itemconfig("content_frame", width=e.width))


# Update scroll region when contentFrame size changes
contentFrame.bind("<Configure>", updateScrollregion)
mainCanvas.bind_all("<MouseWheel>", lambda e: onMousewheel(e, mainCanvas))
mainCanvas.bind_all("<Button-4>", lambda e: onMousewheel(e, mainCanvas))
mainCanvas.bind_all("<Button-5>", lambda e: onMousewheel(e, mainCanvas))

# --- Header Section ---
headerFrame = Frame(contentFrame, bg="#030075")
headerFrame.pack(side="top", fill="x", pady=(10, 5), padx=20)
try:
    logoImg = Image.open("Assets/PNG/Asset 1.png")
    logoImg = logoImg.resize((60, 60))
    logoTk = ImageTk.PhotoImage(logoImg)
    logoLabel = Label(headerFrame, image=logoTk, bg="#030075")
    logoLabel.image = logoTk
    logoLabel.pack(side="left", padx=(0, 20))
except Exception as e:
    print(f"Error loading logo: {e}")
    Label(headerFrame, text="[Logo]", bg="#030075", fg="white").pack(side="left", padx=(0, 20))
titleFrame = Frame(headerFrame, bg="#030075")
titleFrame.pack(side="left", expand=True, fill="x")
titleLabel = Label(titleFrame, text="Weekly Demand Forecasting System", font=("Arial", 20, "bold"), fg="white", bg="#030075")
titleLabel.pack(pady=10)

# --- Control Buttons Section ---
buttonsWrapperFrame = Frame(contentFrame, bg="#F5F5F5")
buttonsWrapperFrame.pack(pady=10, padx=20, fill=tk.X)
buttonsFrame = Frame(buttonsWrapperFrame, bg="#F5F5F5")
buttonsFrame.pack(pady=5)
btnLoadData = Button(buttonsFrame, text="Load Dataset", command=loadDataset, bg="#FFC40A", fg="black", font=("Arial", 10, "bold"), width=15, padx=10, pady=5, cursor="hand2", relief=tk.RAISED)
btnLoadData.pack(side=tk.LEFT, padx=10)
btnProcessDataset = Button(buttonsFrame, text="Process Dataset", command=processDatasetAndShowSelector, bg="#FFC40A", fg="black", font=("Arial", 10, "bold"), width=15, padx=10, pady=5, cursor="hand2", relief=tk.RAISED, state=tk.DISABLED)
btnProcessDataset.pack(side=tk.LEFT, padx=10)
btnGenerateForecast = Button(buttonsFrame, text="Generate Forecast", command=callForecasting, bg="#2ECC71", fg="white", font=("Arial", 10, "bold"), width=18, padx=10, pady=5, cursor="hand2", relief=tk.RAISED, state=tk.DISABLED)
btnGenerateForecast.pack(side=tk.LEFT, padx=10)

# --- Results Text Area Section ---
resultOuterFrame = Frame(contentFrame, bg="#5B53FF")
resultOuterFrame.pack(pady=10, padx=20)
Label(resultOuterFrame, text="Processing Log & Results", font=("Arial", 11, "bold"), fg="white", bg="#5B53FF").pack(anchor="w", padx=10, pady=(5, 2))
resultDisplayFrame = Frame(resultOuterFrame, bd=1, relief=tk.SOLID, bg="white")
resultDisplayFrame.pack(padx=5, pady=(0, 5))
resultText = Text(resultDisplayFrame, width=50, height=4, bd=0, font=("Consolas", 9), wrap=tk.WORD, relief=tk.FLAT)
resultTextScrollbar = ttk.Scrollbar(resultDisplayFrame, command=resultText.yview)
resultText['yscrollcommand'] = resultTextScrollbar.set
resultTextScrollbar.pack(side=tk.RIGHT, fill=tk.Y)
resultText.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=1, pady=1)
resultText.bind("<MouseWheel>", lambda e: onMousewheel(e, resultText))
resultText.bind("<Button-4>", lambda e: onMousewheel(e, resultText))
resultText.bind("<Button-5>", lambda e: onMousewheel(e, resultText))

# --- Main Data Display Area (Table and Chart) ---
mainDisplayFrame = Frame(contentFrame, bg="#F0F0F0")
mainDisplayFrame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

# --- Data Table Section (Left) ---
dataTableOuterFrame = Frame(mainDisplayFrame, bg="#5B53FF")
dataTableOuterFrame.pack(side=tk.LEFT, padx=(0, 10), fill=tk.BOTH, expand=True)
Label(dataTableOuterFrame, text="Selected Products Overview", font=("Arial", 11, "bold"), fg="white", bg="#5B53FF").pack(anchor="w", padx=10, pady=(5, 2))
dataTableFrame = Frame(dataTableOuterFrame, bd=1, relief=tk.SOLID, bg="white")
dataTableFrame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
dataTableScrollbar = ttk.Scrollbar(dataTableFrame)
dataTableScrollbar.pack(side=tk.RIGHT, fill=tk.Y)
# Define Treeview style here if not done in selection window
style = ttk.Style(window) # Use main window for style definition
style.configure("Treeview", rowheight=25, font=("Arial", 9))
style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
style.map("Treeview", background=[("selected", "#a6d1ff")])
dataTable = ttk.Treeview(dataTableFrame, yscrollcommand=dataTableScrollbar.set, height=12, show="headings", style="Treeview")
dataTableScrollbar.config(command=dataTable.yview)
dataTable.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
dataTable.bind("<MouseWheel>", lambda e: onMousewheel(e, dataTable))
dataTable.bind("<Button-4>", lambda e: onMousewheel(e, dataTable))
dataTable.bind("<Button-5>", lambda e: onMousewheel(e, dataTable))

# --- Chart Section (Right) ---
chartOuterFrame = Frame(mainDisplayFrame, bg="#5B53FF")
chartOuterFrame.pack(side=tk.LEFT, padx=(10, 0), fill=tk.BOTH, expand=True)
Label(chartOuterFrame, text="Weekly Demand Chart", font=("Arial", 11, "bold"), fg="white", bg="#5B53FF").pack(anchor="w", padx=10, pady=(5, 2)) # Updated label
chartContainerFrame = Frame(chartOuterFrame, bd=1, relief=tk.SOLID, bg="white")
chartContainerFrame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
chartDisplayFrame = Frame(chartContainerFrame, bg="white")
chartDisplayFrame.pack(fill=tk.BOTH, expand=True)

Button(contentFrame, text="Enviar correo", command=abrir_correo,
       bg="blue", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

# --- LPM Solution Section ---
lpmOuterFrame = Frame(contentFrame, bg="#5B53FF") # Blue accent
lpmOuterFrame.pack(pady=15, padx=20)

Label(lpmOuterFrame, text="LPM Solution Panel",
       font=("Arial", 12, "bold"), fg="white", bg="#5B53FF").pack(pady=5)

lpmInnerFrame = Frame(lpmOuterFrame, bd=1, relief=tk.SOLID, bg="white")
lpmInnerFrame.pack(padx=5, pady=(0, 5), fill=tk.BOTH, expand=True)
lpmOptionsFrame = Frame(lpmInnerFrame, bg="white", width=150)
lpmOptionsFrame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)
lpmOptionsFrame.pack_propagate(False)
lpmInputFrame = Frame(lpmInnerFrame, bg="white")
lpmInputFrame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
lpmOptionVar = StringVar(value="balance")
Radiobutton(lpmOptionsFrame, text="Balance Calculation", variable=lpmOptionVar, value="balance", bg="white", anchor="w", command=updateLpmPanel, font=("Arial", 9)).pack(anchor="w", pady=3, fill="x")
Radiobutton(lpmOptionsFrame, text="Option 2", variable=lpmOptionVar, value="option2", bg="white", anchor="w", command=updateLpmPanel, font=("Arial", 9)).pack(anchor="w", pady=3, fill="x")
Radiobutton(lpmOptionsFrame, text="Option 3", variable=lpmOptionVar, value="option3", bg="white", anchor="w", command=updateLpmPanel, font=("Arial", 9)).pack(anchor="w", pady=3, fill="x")

# --- Footer Section ---
footerFrame = Frame(contentFrame, bg="#E0E0E0", height=30)
footerFrame.pack(fill=tk.X, side=tk.BOTTOM, pady=(15, 0), padx=20)
footerFrame.pack_propagate(False)
footerLabel = Label(footerFrame, text="MICI - Weekly Demand Forecasting System v2.1 | Powered by Gemini Pro", font=("Arial", 8), fg="#555", bg="#E0E0E0") # Updated footer text
footerLabel.pack(pady=5)

# --- Final Initialization ---
updateLpmPanel()
resetUIState()
window.bind("<Configure>", updateScrollregion)

window.mainloop()
